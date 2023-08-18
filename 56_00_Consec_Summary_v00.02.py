import os
import sys
import glob
import time
import openpyxl

from datetime import datetime
Script_Start = datetime.now()
Author= "Antl"
Modification_Date =  "2021.02.18"
Code_Ver = "00.02"


Script_Path = os.path.abspath(__file__)
Script_Directory = os.path.dirname(Script_Path)
os.chdir(Script_Directory)

Quit_Delay = 3
FNULL = open(os.devnull, "w")
Code_File_Ver = os.path.basename(__file__).split("_v")[1].split("_Egemen")[0].replace(".py","")
if not Code_Ver == Code_File_Ver:
    print(str(datetime.now() - Script_Start) + " !Error-001 : Version does not match.")
    time.sleep(Quit_Delay)
    os.remove(os.path.basename(__file__))
    sys.exit()
if datetime.now().year > 2021:
    print(str(datetime.now() - Script_Start) + " !Error-002 : Licence not found")
    time.sleep(Quit_Delay)
    os.remove(os.path.basename(__file__))
    sys.exit()

DB = {}
LC = "N/A"

for File in glob.iglob(Script_Directory + "\\**\\**.res", recursive=True):
    SLS_Switch = "ULS/ALS"
    Read_Switch_ULS_Steel = False
    Read_Switch_Concrete = False
    Read_Switch_SLS_Crack = False
    Read_Switch_SLS_Crack_Next = False
    Thermal_Active = False
    with open(File, 'r' , encoding='ansi') as File_In:
        Lines = File_In.readlines()
        for Line in Lines:
            if "THERMAL EFFECTS :" in Line:
                Thermal_Active = True

            if "SLS" in Line:
                SLS_Switch = "SLS"

            if "LC:" in Line:
                Temp_LC =  Line.split("LC:")[1].strip().split()[0].replace("Â","")
                if Thermal_Active:
                    if not "Y" in Temp_LC:
                        Temp_LC = "T" + Temp_LC

            if "Element :" in Line:
                Wall = Line.split("Element :")[1].split()[0]
                if not Wall in DB:
                    DB[Wall]= {}

            if "Node :" in Line:
                Node = Line.split("Node :")[1].split()[0]
                if not Node in DB[Wall]:
                    DB[Wall][Node]={}
                    DB[Wall][Node]["ULS/ALS"] = {}
                    DB[Wall][Node]["SLS"] = {}
                if not Temp_LC in DB[Wall][Node][SLS_Switch]:
                    DB[Wall][Node][SLS_Switch][Temp_LC] = {}
                    DB[Wall][Node][SLS_Switch][Temp_LC]["Concrete"] = {}
                    DB[Wall][Node][SLS_Switch][Temp_LC]["Concrete"]["Min"] = 100000
                    DB[Wall][Node][SLS_Switch][Temp_LC]["Concrete"]["Max"] = -100000
                    DB[Wall][Node][SLS_Switch][Temp_LC]["Crack"] = {}
                    DB[Wall][Node][SLS_Switch][Temp_LC]["Crack"]["Top"] = {}
                    if not  SLS_Switch == "SLS":
                        DB[Wall][Node][SLS_Switch][Temp_LC]["Steel"] = {}
                        DB[Wall][Node][SLS_Switch][Temp_LC]["Steel"]["Strain"] = {}
                        DB[Wall][Node][SLS_Switch][Temp_LC]["Steel"]["Stress"] = {}


            if Read_Switch_SLS_Crack:
                if not Read_Switch_SLS_Crack_Next:
                    DB[Wall][Node][SLS_Switch][Temp_LC]["Crack"]["Top"] = float(Line.strip()[29:37])
                else:
                    DB[Wall][Node][SLS_Switch][Temp_LC]["Crack"]["Bot"] = float(Line.strip()[29:37])

                if Read_Switch_SLS_Crack_Next:
                    Read_Switch_SLS_Crack = False
                Read_Switch_SLS_Crack_Next = True


            if SLS_Switch == "SLS":
                if "wok" in Line:
                    Read_Switch_SLS_Crack = True
                    Read_Switch_SLS_Crack_Next = False
                    

            if "steel" in Line:
                Read_Switch_Concrete = False
                Thermal_Active = False
            if Read_Switch_Concrete:
                Text = Line
                if "z+" in Line:
                    Text = Line.replace("z+","")
                if "z-" in Line:
                    Text = Line.replace("z-","")
                for i in range(0,2):
                    Temp_Stress = float(Text.split()[3 + i])
                    if Temp_Stress > DB[Wall][Node][SLS_Switch][Temp_LC]["Concrete"]["Max"]:
                        DB[Wall][Node][SLS_Switch][Temp_LC]["Concrete"]["Max"] = Temp_Stress
                    if Temp_Stress < DB[Wall][Node][SLS_Switch][Temp_LC]["Concrete"]["Min"]:
                        DB[Wall][Node][SLS_Switch][Temp_LC]["Concrete"]["Min"] = Temp_Stress
            if not SLS_Switch == "SLS":
                if "layer          e1         e2       s1       s2      angle" in Line:
                    Read_Switch_Concrete = True

            if Read_Switch_ULS_Steel:
                try:
                    Stress = float(Line.split()[2])
                    Layer = int(Line.split()[0])
                    Strain = float(Line.split()[1])
                    DB[Wall][Node][SLS_Switch][Temp_LC]["Steel"]["Strain"]["L" + str(Layer)] = Strain
                    DB[Wall][Node][SLS_Switch][Temp_LC]["Steel"]["Stress"]["L" + str(Layer)] = Stress
                except:
                    Read_Switch_ULS_Steel = False

            if not  SLS_Switch == "SLS":
                if "layer      strain     stress    angle       zb         as" in Line:
                    Read_Switch_ULS_Steel = True




WorkBook_1 = openpyxl.Workbook()
WorkSheet_1 = WorkBook_1.active
WorkSheet_1.title = "ULS ALS"
Row_Number = 1
WorkSheet_1.cell(column = 2, row = Row_Number, value = "Member")
WorkSheet_1.cell(column = 3, row = Row_Number, value = "Node")
WorkSheet_1.cell(column = 4, row = Row_Number, value = "LC")
WorkSheet_1.cell(column = 5, row = Row_Number, value = "smax")
WorkSheet_1.cell(column = 6, row = Row_Number, value = "smin")
for Wall in DB:
    for Node in DB[Wall]:
        for LC in DB[Wall][Node]["ULS/ALS"]:
            Row_Number = Row_Number + 1
            WorkSheet_1.cell(column = 2, row = Row_Number, value = Wall)
            WorkSheet_1.cell(column = 3, row = Row_Number, value = Node)
            WorkSheet_1.cell(column = 4, row = Row_Number, value = LC)
            WorkSheet_1.cell(column = 5, row = Row_Number, value = DB[Wall][Node]["ULS/ALS"][LC]["Concrete"]["Max"])
            WorkSheet_1.cell(column = 6, row = Row_Number, value = DB[Wall][Node]["ULS/ALS"][LC]["Concrete"]["Min"])
            i = 0
            for Layer in DB[Wall][Node]["ULS/ALS"][LC]["Steel"]["Stress"]:
                WorkSheet_1.cell(column = 7 + i, row = Row_Number, value = DB[Wall][Node]["ULS/ALS"][LC]["Steel"]["Stress"][Layer])
                if abs(DB[Wall][Node]["ULS/ALS"][LC]["Steel"]["Strain"][Layer]) > 2.5:
                    WorkSheet_1.cell(column = 11 + i, row = Row_Number, value = 1)
                else:
                    UR = round(abs(DB[Wall][Node]["ULS/ALS"][LC]["Steel"]["Strain"][Layer])/2.5,2)
                    WorkSheet_1.cell(column = 11 + i, row = Row_Number, value = UR)
                i = i + 1
WorkBook_1.save(filename = "Summary_ULS.xlsx")

WorkBook_2 = openpyxl.Workbook()
WorkSheet_2 = WorkBook_2.active

WorkSheet_2.title = "SLS"
Row_Number = 1
WorkSheet_2.cell(column = 2, row = Row_Number, value = "Member")
WorkSheet_2.cell(column = 3, row = Row_Number, value = "Node")
WorkSheet_2.cell(column = 4, row = Row_Number, value = "LC")
WorkSheet_2.cell(column = 5, row = Row_Number, value = "wk+")
WorkSheet_2.cell(column = 6, row = Row_Number, value = "wk-")
for Wall in DB:
    for Node in DB[Wall]:
        for LC in DB[Wall][Node]["SLS"]:
            Row_Number = Row_Number + 1
            WorkSheet_2.cell(column = 2, row = Row_Number, value = Wall)
            WorkSheet_2.cell(column = 3, row = Row_Number, value = Node)
            WorkSheet_2.cell(column = 4, row = Row_Number, value = LC)
            WorkSheet_2.cell(column = 5, row = Row_Number, value = DB[Wall][Node]["SLS"][LC]["Crack"]["Top"])
            WorkSheet_2.cell(column = 6, row = Row_Number, value = DB[Wall][Node]["SLS"][LC]["Crack"]["Bot"])

WorkBook_2.save(filename = "Summary_SLS.xlsx")
